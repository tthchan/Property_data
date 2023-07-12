from selenium import webdriver
from bs4 import BeautifulSoup
import pandas as pd
from datetime import date, datetime, timedelta
import openpyxl
from openpyxl import load_workbook

#set webdriver
chrome_options = webdriver.chrome.options.Options()

prefs = {
'download.extensions_to_open': 'xml',
'safebrowsing.enabled': True
}
## Uncomment the below to let the program to run in background
options = webdriver.ChromeOptions()
options.add_experimental_option('prefs',prefs)
options.add_argument("start-maximized")
options.add_argument("--headless")
# options.add_argument("disable-infobars")
options.add_argument("--disable-extensions")
options.add_argument("--safebrowsing-disable-download-protection")
options.add_argument("safebrowsing-disable-extension-blacklist")
options.add_argument('--no-sandbox')
options.add_argument('--disable-dev-shm-usage')

workbook = load_workbook(filename='Centaline_sale_lease.xlsx')
sheet = workbook.active
Date = []
for cell in sheet.iter_cols(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
    for elem in cell:
        Date.append(elem.value.date())
Sale = []
for cell in sheet.iter_cols(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2, values_only=True):
    for elem in cell:
        Sale.append(elem)
Lease = []
for cell in sheet.iter_cols(min_row=2, max_row=sheet.max_row, min_col=3, max_col=3,values_only=True):
    for elem in cell:
        Lease.append(elem)

#Sale
driver = webdriver.Chrome('/usr/bin/chromedriver', options = options)
url = 'https://hk.centanet.com/findproperty/list/buy'
driver.get(url)
sale = driver.find_element_by_xpath('//*[@id="__layout"]/div/div[4]/div[6]/div/div[1]/div[1]/div/h2/span/span').text
print(sale.replace(",",""))
Sale.append(int(sale.replace(",","")))
driver.quit()

#Lease
driver = webdriver.Chrome('/usr/bin/chromedriver', options = options)
url = 'https://hk.centanet.com/findproperty/list/rent'
driver.get(url)
lease = driver.find_element_by_xpath('//*[@id="__layout"]/div/div[4]/div[6]/div/div[1]/div[1]/div/h2/span/span').text
print(lease.replace(",",""))
Lease.append(int(lease.replace(",","")))
driver.quit()

#Date
d = date.today() + timedelta(days=1)
print(d)
Date.append(d)

df=pd.DataFrame()
df["日期"] = Date
df["賣盤"] = Sale
df["租盤"] = Lease
df.to_excel("Centaline_sale_lease.xlsx",sheet_name="sheet1", index=False)
